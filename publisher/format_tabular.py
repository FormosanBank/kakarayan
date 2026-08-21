"""Prepared relational, hierarchical, spreadsheet, and text exports."""

from __future__ import annotations

import csv
import json
import re
import sqlite3
import tempfile
import zipfile
from collections.abc import Iterator
from datetime import UTC, datetime
from itertools import chain
from pathlib import Path
from typing import Any

import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from publisher import SCHEMA_VERSION
from publisher.archive import directory_entries, repack_zip, write_zip
from publisher.release_db import open_release
from publisher.tables import INTEGER_COLUMNS, REAL_COLUMNS, TABLE_COLUMNS

_BATCH_SIZE = 50_000
_XLSX_ROWS_PER_SHEET = 1_000_000


def _json(value: object) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _safe_spreadsheet(value: object) -> object:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@", "\t", "\r")):
        return f"'{value}"
    return value


def _arrow_schema(table: str, release_id: str) -> pa.Schema:
    fields = []
    for column in TABLE_COLUMNS[table]:
        kind: pa.DataType
        if column in INTEGER_COLUMNS:
            kind = pa.int64()
        elif column in REAL_COLUMNS:
            kind = pa.float64()
        else:
            kind = pa.string()
        fields.append(pa.field(column, kind, nullable=True))
    return pa.schema(
        fields,
        metadata={
            b"kakarayan_schema_version": SCHEMA_VERSION.encode(),
            b"kakarayan_release_id": release_id.encode(),
            b"source_representation": b"FormosanBank canonical XML",
        },
    )


def write_tsv(database: Path, output: Path) -> None:
    output.mkdir(parents=True, exist_ok=True)
    with open_release(database) as connection:
        for table, columns in TABLE_COLUMNS.items():
            with (output / f"{table}.tsv").open(
                "w",
                encoding="utf-8",
                newline="",
            ) as stream:
                writer = csv.writer(stream, delimiter="\t", lineterminator="\n")
                writer.writerow(columns)
                cursor = connection.execute(
                    f'SELECT {", ".join(columns)} FROM "{table}" ORDER BY rowid'
                )
                for row in cursor:
                    writer.writerow([r"\N" if value is None else value for value in row])


def write_parquet(database: Path, output: Path, release_id: str) -> dict[str, object]:
    output.mkdir(parents=True, exist_ok=True)
    schemas: dict[str, object] = {}
    with open_release(database) as connection:
        for table, columns in TABLE_COLUMNS.items():
            schema = _arrow_schema(table, release_id)
            path = output / f"{table}.parquet"
            with pq.ParquetWriter(
                path,
                schema,
                compression="zstd",
                use_dictionary=True,
                version="2.6",
                write_statistics=True,
            ) as writer:
                cursor = connection.execute(
                    f'SELECT {", ".join(columns)} FROM "{table}" ORDER BY rowid'
                )
                while rows := cursor.fetchmany(_BATCH_SIZE):
                    arrays = [
                        pa.array([row[index] for row in rows], type=field.type)
                        for index, field in enumerate(schema)
                    ]
                    writer.write_batch(
                        pa.RecordBatch.from_arrays(arrays, schema=schema),
                        row_group_size=_BATCH_SIZE,
                    )
            schemas[table] = {
                "path": f"{table}.parquet",
                "fields": [
                    {
                        "name": field.name,
                        "type": str(field.type),
                        "nullable": field.nullable,
                    }
                    for field in schema
                ],
                "compression": "zstd",
                "row_group_size": _BATCH_SIZE,
            }
    return schemas


def _tiers(
    connection: sqlite3.Connection,
    owner_type: str,
    owner_id: str,
) -> dict[str, list[dict[str, Any]]]:
    return {
        table: [
            dict(row)
            for row in connection.execute(
                f"SELECT * FROM {table} WHERE owner_type = ? AND owner_id = ? ORDER BY position",
                (owner_type, owner_id),
            )
        ]
        for table in ("forms", "phonology", "translations", "audio")
    }


def _sentence_records(
    connection: sqlite3.Connection,
    *,
    partitioned: bool = False,
) -> Iterator[dict[str, Any]]:
    order = (
        "language_id, corpus_id, source_path, source_ordinal, sentence_id"
        if partitioned
        else "source_path, source_ordinal, sentence_id"
    )
    cursor = connection.execute(f"SELECT * FROM sentence_view ORDER BY {order}")
    for sentence_row in cursor:
        sentence = dict(sentence_row)
        sentence_id = sentence["sentence_id"]
        sentence["tiers"] = _tiers(connection, "sentence", sentence_id)
        sentence["tokens"] = [
            dict(row)
            for row in connection.execute(
                "SELECT * FROM tokens WHERE sentence_id = ? ORDER BY position",
                (sentence_id,),
            )
        ]
        words = [
            dict(row)
            for row in connection.execute(
                "SELECT * FROM words WHERE parent_id = ? ORDER BY position",
                (sentence_id,),
            )
        ]
        for word in words:
            word["tiers"] = _tiers(connection, "word", word["id"])
            morphemes = [
                dict(row)
                for row in connection.execute(
                    "SELECT * FROM morphemes WHERE parent_id = ? ORDER BY position",
                    (word["id"],),
                )
            ]
            for morpheme in morphemes:
                morpheme["tiers"] = _tiers(connection, "morpheme", morpheme["id"])
            word["morphemes"] = morphemes
        sentence["words"] = words
        yield sentence


def write_hierarchical_jsonl(
    database: Path,
    output: Path,
    release_id: str,
    *,
    rows_per_part: int = 20_000,
    package_metadata: list[tuple[str, Path]] | None = None,
) -> dict[str, object]:
    output.mkdir(parents=True, exist_ok=True)
    partitions: list[dict[str, object]] = []
    with tempfile.TemporaryDirectory(prefix="kakarayan-jsonl-") as temporary_name:
        temporary = Path(temporary_name)
        scope: tuple[str, str] | None = None
        stream: Any = None
        part = 0
        part_rows = 0
        scope_rows = 0

        def close_stream() -> None:
            nonlocal stream
            if stream is not None:
                stream.close()
                stream = None

        def package_scope() -> None:
            nonlocal part, part_rows, scope_rows
            close_stream()
            if scope is None:
                return
            language_id, corpus_id = scope
            zip_path = output / f"{language_id}--{corpus_id}.zip"
            write_zip(
                zip_path,
                chain(directory_entries(temporary), package_metadata or []),
            )
            partitions.append(
                {
                    "path": f"jsonl/{zip_path.name}",
                    "language_id": language_id,
                    "corpus_id": corpus_id,
                    "records": scope_rows,
                    "parts": part + int(part_rows > 0),
                    "bytes": zip_path.stat().st_size,
                }
            )
            for candidate in temporary.iterdir():
                candidate.unlink()
            part = 0
            part_rows = 0
            scope_rows = 0

        with open_release(database) as connection:
            for sentence in _sentence_records(connection, partitioned=True):
                next_scope = (sentence["language_id"], sentence["corpus_id"])
                if next_scope != scope:
                    package_scope()
                    scope = next_scope
                if stream is None or part_rows >= rows_per_part:
                    close_stream()
                    if part_rows:
                        part += 1
                    part_rows = 0
                    stream = (temporary / f"part-{part:04d}.jsonl").open(
                        "w",
                        encoding="utf-8",
                        newline="\n",
                    )
                stream.write(_json(sentence) + "\n")
                part_rows += 1
                scope_rows += 1
            package_scope()
    return {
        "release_id": release_id,
        "record_unit": "hierarchical sentence",
        "rows_per_part": rows_per_part,
        "partitions": partitions,
    }


def write_plain_text(database: Path, output: Path) -> None:
    output.mkdir(parents=True, exist_ok=True)
    sentence_query = """
        SELECT sentence_id, language_id, corpus_id, source_path, source_ordinal,
               COALESCE(standard_form, original_form, '') AS source_form,
               COALESCE((
                 SELECT group_concat(text, ' | ')
                 FROM (
                   SELECT text FROM translations
                   WHERE owner_type = 'sentence' AND owner_id = sentence_id
                   ORDER BY position
                 )
               ), '') AS translation
        FROM sentence_view
        ORDER BY source_path, source_ordinal, sentence_id
    """
    with (
        open_release(database) as connection,
        (output / "sentences.txt").open("w", encoding="utf-8", newline="\n") as source,
        (output / "parallel.tsv").open("w", encoding="utf-8", newline="") as parallel,
        (output / "interlinear.txt").open("w", encoding="utf-8", newline="\n") as interlinear,
    ):
        writer = csv.writer(parallel, delimiter="\t", lineterminator="\n")
        writer.writerow(
            [
                "sentence_id",
                "language_id",
                "corpus_id",
                "source_path",
                "source_ordinal",
                "source_form",
                "translation",
            ]
        )
        rows = connection.execute(sentence_query)
        for row in rows:
            source.write(f"{row['source_form']}\n")
            writer.writerow(row)
            token_line = " ".join(
                item[0]
                for item in connection.execute(
                    "SELECT surface FROM tokens WHERE sentence_id = ? ORDER BY position",
                    (row["sentence_id"],),
                )
            )
            interlinear.write(
                f"\\id {row['sentence_id']}\n"
                f"\\tx {row['source_form']}\n"
                f"\\mb {token_line}\n"
                f"\\ft {row['translation']}\n\n"
            )


def write_audio_manifest(database: Path, output: Path) -> None:
    query = """
        WITH owner_scope(owner_type, owner_id, text_id, sentence_id) AS (
          SELECT 'text', id, id, NULL FROM texts
          UNION ALL
          SELECT 'sentence', id, parent_id, id FROM sentences
          UNION ALL
          SELECT 'word', w.id, s.parent_id, s.id
          FROM words w JOIN sentences s ON s.id = w.parent_id
          UNION ALL
          SELECT 'morpheme', m.id, s.parent_id, s.id
          FROM morphemes m
          JOIN words w ON w.id = m.parent_id
          JOIN sentences s ON s.id = w.parent_id
        )
        SELECT a.*, scope.text_id, scope.sentence_id, t.corpus_id, t.language_id,
               t.dialect, t.source_path
        FROM audio a
        JOIN owner_scope scope
          ON scope.owner_type = a.owner_type AND scope.owner_id = a.owner_id
        JOIN texts t ON t.id = scope.text_id
        ORDER BY t.source_path, a.owner_type, a.owner_id, a.position
    """
    with (
        open_release(database) as connection,
        output.open(
            "w",
            encoding="utf-8",
            newline="",
        ) as stream,
    ):
        cursor = connection.execute(query)
        columns = tuple(item[0] for item in cursor.description)
        writer = csv.writer(stream, delimiter="\t", lineterminator="\n")
        writer.writerow(columns)
        for row in cursor:
            writer.writerow([r"\N" if value is None else value for value in row])


def write_xlsx(
    database: Path,
    path: Path,
    release_id: str,
    source_commit: str,
    rights: dict[str, object],
) -> None:
    raw = path.with_suffix(".raw.xlsx")
    workbook = Workbook(write_only=True)
    workbook.properties.created = datetime(1980, 1, 1, tzinfo=UTC)
    workbook.properties.modified = datetime(1980, 1, 1, tzinfo=UTC)
    readme = workbook.create_sheet("README")
    for row in (
        ("Kakarayan FormosanBank research workbook",),
        ("release_id", release_id),
        ("source_commit", source_commit),
        ("schema_version", SCHEMA_VERSION),
        ("null representation", "blank cell"),
        ("formula safety", "text beginning = + - @ tab or CR is prefixed with apostrophe"),
        ("canonical representation", "Use the pinned FormosanBank XML for archival work."),
        ("rights", "See the RIGHTS sheet. Public visibility does not imply uniform reuse."),
    ):
        readme.append(row)
    rights_sheet = workbook.create_sheet("RIGHTS")
    rights_columns = (
        "id",
        "corpus",
        "review_status",
        "redistribution",
        "commercial_use",
        "derivatives",
        "notice",
    )
    rights_sheet.append(rights_columns)
    rights_entries = rights.get("entries")
    if not isinstance(rights_entries, list):
        raise ValueError("Rights metadata has no entries")
    for entry in rights_entries:
        if not isinstance(entry, dict):
            continue
        rights_sheet.append([_safe_spreadsheet(entry.get(column)) for column in rights_columns])
    with open_release(database) as connection:
        for table, columns in TABLE_COLUMNS.items():
            cursor = connection.execute(
                f'SELECT {", ".join(columns)} FROM "{table}" ORDER BY rowid'
            )
            part = 1
            row_number = 0
            worksheet = None
            for row in cursor:
                if worksheet is None or row_number >= _XLSX_ROWS_PER_SHEET:
                    title = table[:25] if part == 1 else f"{table[:25]}_{part}"
                    worksheet = workbook.create_sheet(title)
                    worksheet.freeze_panes = "A2"
                    worksheet.auto_filter.ref = (
                        f"A1:{get_column_letter(len(columns))}{_XLSX_ROWS_PER_SHEET + 1}"
                    )
                    for index, column in enumerate(columns, 1):
                        worksheet.column_dimensions[get_column_letter(index)].width = min(
                            max(len(column) + 2, 12),
                            36,
                        )
                    worksheet.append(columns)
                    part += 1
                    row_number = 0
                worksheet.append([_safe_spreadsheet(value) for value in row])
                row_number += 1
            if worksheet is None:
                worksheet = workbook.create_sheet(table[:25])
                worksheet.freeze_panes = "A2"
                worksheet.append(columns)
    workbook.save(raw)
    with zipfile.ZipFile(raw) as source:
        core = re.sub(
            rb"(<dcterms:modified[^>]*>).*?(</dcterms:modified>)",
            rb"\g<1>1980-01-01T00:00:00Z\g<2>",
            source.read("docProps/core.xml"),
        )
    repack_zip(path, raw, replacements={"docProps/core.xml": core})
    raw.unlink()


def data_dictionary() -> dict[str, object]:
    return {
        "schema_version": SCHEMA_VERSION,
        "null_representation": {
            "csv_tsv": r"\N",
            "json": None,
            "sqlite": "NULL",
            "xlsx": "blank cell",
        },
        "tables": {
            table: [
                {
                    "name": column,
                    "logical_type": (
                        "integer"
                        if column in INTEGER_COLUMNS
                        else "number"
                        if column in REAL_COLUMNS
                        else "string"
                    ),
                }
                for column in columns
            ]
            for table, columns in TABLE_COLUMNS.items()
        },
        "tier_ownership": {
            "sentence": "A translation with owner_type=sentence applies only to owner_id sentence.",
            "word": "A translation with owner_type=word applies only to owner_id word.",
            "morpheme": (
                "A translation with owner_type=morpheme applies only to owner_id morpheme."
            ),
            "joins": (
                "Join translations on both owner_type and owner_id. "
                "Never infer ownership from row order."
            ),
        },
    }
