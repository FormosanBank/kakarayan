"""Reconcile counts and deterministic samples across release representations."""

from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import io
import json
import math
import shutil
import sqlite3
import tempfile
import zipfile
from collections.abc import Iterator, Mapping
from contextlib import closing, contextmanager
from pathlib import Path
from typing import Any

import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook

from publisher.tables import INTEGER_COLUMNS, REAL_COLUMNS, TABLE_COLUMNS
from publisher.verify_release import VerificationError, verify_release

_DELIMITED_FIELD_LIMIT = 64 * 1024 * 1024


class ReconciliationError(RuntimeError):
    """Raised when two generated representations disagree."""


def _require_equal(label: str, actual: object, expected: object) -> None:
    if actual != expected:
        raise ReconciliationError(f"{label} differs: expected {expected!r}, received {actual!r}")


@contextmanager
def _database(root: Path) -> Iterator[Path]:
    direct = root / "formosanbank.sqlite"
    if direct.is_file():
        yield direct
        return
    compressed = root / "formosanbank.sqlite.gz"
    if not compressed.is_file():
        raise ReconciliationError("Release has no SQLite representation")
    with tempfile.NamedTemporaryFile(prefix="kakarayan-reconcile-", suffix=".sqlite") as output:
        with gzip.open(compressed, "rb") as source:
            shutil.copyfileobj(source, output, length=1024 * 1024)
        output.flush()
        yield Path(output.name)


def _database_state(
    database: Path,
) -> tuple[
    dict[str, int],
    dict[str, dict[str, Any]],
    float,
    dict[str, int],
    float,
]:
    counts: dict[str, int] = {}
    samples: dict[str, dict[str, Any]] = {}
    uri = f"file:{database}?mode=ro&immutable=1"
    with closing(sqlite3.connect(uri, uri=True)) as connection:
        connection.row_factory = sqlite3.Row
        if connection.execute("PRAGMA integrity_check").fetchone()[0] != "ok":
            raise ReconciliationError("SQLite integrity check failed")
        for table, columns in TABLE_COLUMNS.items():
            counts[table] = int(connection.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0])
            row = connection.execute(
                f'SELECT {", ".join(columns)} FROM "{table}" ORDER BY id LIMIT 1'
            ).fetchone()
            if row is not None:
                samples[table] = dict(row)
        duration = math.fsum(
            float(row[0])
            for row in connection.execute("SELECT duration FROM audio WHERE duration IS NOT NULL")
        )
        hierarchical_counts = dict(counts)
        for table in ("forms", "phonology", "translations", "audio"):
            hierarchical_counts[table] = int(
                connection.execute(
                    f'SELECT COUNT(*) FROM "{table}" WHERE owner_type != ?',
                    ("text",),
                ).fetchone()[0]
            )
        hierarchical_duration = math.fsum(
            float(row[0])
            for row in connection.execute(
                "SELECT duration FROM audio WHERE duration IS NOT NULL AND owner_type != ?",
                ("text",),
            )
        )
    return counts, samples, duration, hierarchical_counts, hierarchical_duration


def _coerce_row(table: str, row: Mapping[str, Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for column in TABLE_COLUMNS[table]:
        value = row.get(column)
        if value == r"\N":
            value = None
        elif value is not None and column in INTEGER_COLUMNS:
            value = (
                1
                if value is True or value == "True"
                else 0
                if value is False or value == "False"
                else int(value)
            )
        elif value is not None and column in REAL_COLUMNS:
            value = float(value)
        result[column] = value
    return result


def _delimited_archive(
    path: Path,
    suffix: str,
    *,
    delimiter: str,
    samples: Mapping[str, Mapping[str, Any]],
) -> tuple[dict[str, int], dict[str, dict[str, Any]], float]:
    counts: dict[str, int] = {}
    found: dict[str, dict[str, Any]] = {}
    durations: list[float] = []
    previous_limit = csv.field_size_limit()
    csv.field_size_limit(max(previous_limit, _DELIMITED_FIELD_LIMIT))
    try:
        with zipfile.ZipFile(path) as archive:
            for table in TABLE_COLUMNS:
                name = f"{table}.{suffix}"
                if name not in archive.namelist():
                    raise ReconciliationError(f"{path.name} has no {name}")
                with (
                    archive.open(name) as raw,
                    io.TextIOWrapper(raw, encoding="utf-8", newline="") as stream,
                ):
                    reader = csv.DictReader(stream, delimiter=delimiter)
                    _require_equal(
                        f"{path.name}:{name} header",
                        tuple(reader.fieldnames or ()),
                        TABLE_COLUMNS[table],
                    )
                    count = 0
                    sample_id = samples.get(table, {}).get("id")
                    for row in reader:
                        count += 1
                        if row.get("id") == sample_id:
                            found[table] = _coerce_row(table, row)
                        if table == "audio" and row.get("duration") not in {
                            None,
                            "",
                            r"\N",
                        }:
                            durations.append(float(row["duration"]))
                    counts[table] = count
    except csv.Error as error:
        raise ReconciliationError(
            f"{path.name} has a delimited field above {_DELIMITED_FIELD_LIMIT} characters"
        ) from error
    finally:
        csv.field_size_limit(previous_limit)
    return counts, found, math.fsum(durations)


def _jsonl_archive(
    path: Path,
    samples: Mapping[str, Mapping[str, Any]],
) -> tuple[dict[str, int], dict[str, dict[str, Any]], float]:
    counts: dict[str, int] = {}
    found: dict[str, dict[str, Any]] = {}
    durations: list[float] = []
    with zipfile.ZipFile(path) as archive:
        for table in TABLE_COLUMNS:
            name = f"{table}.jsonl"
            if name not in archive.namelist():
                raise ReconciliationError(f"{path.name} has no {name}")
            count = 0
            sample_id = samples.get(table, {}).get("id")
            with archive.open(name) as stream:
                for line in stream:
                    row = json.loads(line)
                    count += 1
                    if row.get("id") == sample_id:
                        found[table] = _coerce_row(table, row)
                    if table == "audio" and row.get("duration") is not None:
                        durations.append(float(row["duration"]))
            counts[table] = count
    return counts, found, math.fsum(durations)


def _parquet_archive(
    path: Path,
    samples: Mapping[str, Mapping[str, Any]],
) -> tuple[dict[str, int], dict[str, dict[str, Any]], float]:
    counts: dict[str, int] = {}
    found: dict[str, dict[str, Any]] = {}
    duration = 0.0
    with zipfile.ZipFile(path) as archive:
        for table, columns in TABLE_COLUMNS.items():
            name = f"{table}.parquet"
            if name not in archive.namelist():
                raise ReconciliationError(f"{path.name} has no {name}")
            data = archive.read(name)
            parquet = pq.ParquetFile(pa.BufferReader(data))
            counts[table] = parquet.metadata.num_rows
            _require_equal(
                f"{path.name}:{name} fields",
                tuple(parquet.schema_arrow.names),
                columns,
            )
            sample_id = samples.get(table, {}).get("id")
            if sample_id is not None:
                rows = pq.read_table(
                    pa.BufferReader(data),
                    filters=[("id", "=", sample_id)],
                ).to_pylist()
                if rows:
                    found[table] = _coerce_row(table, rows[0])
            if table == "audio":
                values = pq.read_table(pa.BufferReader(data), columns=["duration"])["duration"]
                duration = math.fsum(float(value.as_py()) for value in values if value.is_valid)
    return counts, found, duration


def _xlsx_counts(path: Path) -> dict[str, int]:
    counts = {table: 0 for table in TABLE_COLUMNS}
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for table in TABLE_COLUMNS:
            sheets = [
                sheet
                for sheet in workbook.worksheets
                if sheet.title == table or sheet.title.startswith(f"{table}_")
            ]
            if not sheets:
                raise ReconciliationError(f"{path.name} has no {table} sheet")
            for sheet in sheets:
                if sheet.max_row is None:
                    counts[table] += max(sum(1 for _ in sheet.iter_rows(values_only=True)) - 1, 0)
                else:
                    counts[table] += max(sheet.max_row - 1, 0)
    finally:
        workbook.close()
    return counts


def _hierarchical_counts(root: Path) -> tuple[dict[str, int], float]:
    counts = {table: 0 for table in TABLE_COLUMNS}
    text_ids: set[str] = set()
    durations: list[float] = []

    def add_tiers(tiers: Mapping[str, list[dict[str, Any]]]) -> None:
        for table in ("forms", "phonology", "translations", "audio"):
            rows = tiers.get(table, [])
            counts[table] += len(rows)
            if table == "audio":
                durations.extend(
                    float(row["duration"]) for row in rows if row.get("duration") is not None
                )

    for path in sorted((root / "prepared" / "jsonl").glob("*.zip")):
        with zipfile.ZipFile(path) as archive:
            for name in sorted(item for item in archive.namelist() if item.endswith(".jsonl")):
                with archive.open(name) as stream:
                    for line in stream:
                        sentence = json.loads(line)
                        counts["sentences"] += 1
                        text_ids.add(str(sentence["text_id"]))
                        counts["tokens"] += len(sentence["tokens"])
                        add_tiers(sentence["tiers"])
                        for word in sentence["words"]:
                            counts["words"] += 1
                            add_tiers(word["tiers"])
                            for morpheme in word["morphemes"]:
                                counts["morphemes"] += 1
                                add_tiers(morpheme["tiers"])
    counts["texts"] = len(text_ids)
    return counts, math.fsum(durations)


def _canonical_xml(root: Path, source_repo: Path) -> int:
    verified = 0
    for path in sorted((root / "prepared" / "canonical").glob("*.zip")):
        with zipfile.ZipFile(path) as archive:
            manifest = json.loads(archive.read("manifest.json"))
            for entry in manifest["files"]:
                relative = str(entry["path"])
                source = source_repo / relative
                if not source.is_file():
                    raise ReconciliationError(f"Canonical source is missing: {relative}")
                source_digest = hashlib.sha256(source.read_bytes()).hexdigest()
                archive_digest = hashlib.sha256(archive.read(relative)).hexdigest()
                _require_equal(f"canonical manifest {relative}", source_digest, entry["sha256"])
                _require_equal(f"canonical package {relative}", archive_digest, source_digest)
                verified += 1
    return verified


def _browser_counts(site: Path, expected: Mapping[str, int]) -> dict[str, int]:
    response = json.loads(
        (site / "api" / "v1" / "search" / "manifest.json").read_text(encoding="utf-8")
    )
    manifest = response["data"]
    sentences = 0
    tokens = 0
    for shard in manifest["shards"]:
        content = gzip.decompress((site / "data" / shard["path"]).read_bytes())
        records = json.loads(content)
        sentences += len(records)
        tokens += sum(len(record["tokens"]) for record in records)
    result = {"sentences": sentences, "tokens": tokens}
    _require_equal("browser sentence count", sentences, expected["sentences"])
    _require_equal("browser token count", tokens, expected["tokens"])
    return result


def reconcile_release(
    root: Path,
    *,
    source_repo: Path | None = None,
    site: Path | None = None,
) -> dict[str, object]:
    """Validate all primary projections and return a machine-readable report."""
    root = root.resolve()
    try:
        manifest = verify_release(root)
    except VerificationError as error:
        raise ReconciliationError(str(error)) from error
    with _database(root) as database:
        (
            database_counts,
            samples,
            database_duration,
            hierarchical_expected,
            hierarchical_expected_duration,
        ) = _database_state(database)
    _require_equal("release manifest counts", manifest["counts"], database_counts)

    representations: dict[str, dict[str, int]] = {"sqlite": database_counts}
    sample_sets: dict[str, dict[str, dict[str, Any]]] = {}
    duration_totals: dict[str, float] = {"sqlite": database_duration}
    for name, path, suffix, delimiter in (
        ("csv", root / "prepared" / "csv-tables.zip", "csv", ","),
        ("tsv", root / "prepared" / "tsv-tables.zip", "tsv", "\t"),
    ):
        counts, found, duration = _delimited_archive(
            path,
            suffix,
            delimiter=delimiter,
            samples=samples,
        )
        representations[name] = counts
        sample_sets[name] = found
        duration_totals[name] = duration
    jsonl_counts, jsonl_samples, jsonl_duration = _jsonl_archive(
        root / "prepared" / "flat-jsonl-tables.zip",
        samples,
    )
    representations["flat_jsonl"] = jsonl_counts
    sample_sets["flat_jsonl"] = jsonl_samples
    duration_totals["flat_jsonl"] = jsonl_duration
    parquet_counts, parquet_samples, parquet_duration = _parquet_archive(
        root / "prepared" / "parquet-tables.zip",
        samples,
    )
    representations["parquet"] = parquet_counts
    sample_sets["parquet"] = parquet_samples
    duration_totals["parquet"] = parquet_duration
    representations["xlsx"] = _xlsx_counts(root / "prepared" / "formosanbank.xlsx")
    hierarchical_counts, hierarchical_duration = _hierarchical_counts(root)
    representations["hierarchical_jsonl"] = hierarchical_counts
    duration_totals["hierarchical_jsonl"] = hierarchical_duration

    for name, counts in representations.items():
        if name == "hierarchical_jsonl":
            continue
        _require_equal(f"{name} counts", counts, database_counts)
    _require_equal(
        "hierarchical_jsonl counts",
        representations["hierarchical_jsonl"],
        hierarchical_expected,
    )
    for representation, found in sample_sets.items():
        _require_equal(f"{representation} sampled tables", set(found), set(samples))
        for table, row in found.items():
            _require_equal(f"{representation} sample {table}", row, samples[table])
    for name, value in duration_totals.items():
        expected_duration = (
            hierarchical_expected_duration if name == "hierarchical_jsonl" else database_duration
        )
        if not math.isclose(value, expected_duration, rel_tol=1e-12, abs_tol=1e-6):
            raise ReconciliationError(
                f"{name} duration total differs: expected {expected_duration}, received {value}"
            )

    canonical_files = _canonical_xml(root, source_repo.resolve()) if source_repo else None
    browser = _browser_counts(site.resolve(), database_counts) if site else None
    return {
        "release_id": manifest["release_id"],
        "counts": database_counts,
        "duration_seconds": database_duration,
        "representations": sorted(representations),
        "hierarchical_exclusions": [
            "text-owned forms",
            "text-owned phonology",
            "text-owned translations",
            "text-owned audio",
        ],
        "sample_ids": {table: row["id"] for table, row in samples.items()},
        "canonical_files_verified": canonical_files,
        "browser": browser,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--release", required=True, type=Path)
    parser.add_argument("--source-repo", type=Path)
    parser.add_argument("--site", type=Path)
    args = parser.parse_args(argv)
    result = reconcile_release(
        args.release,
        source_repo=args.source_repo,
        site=args.site,
    )
    print(json.dumps(result, ensure_ascii=False, sort_keys=True))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except ReconciliationError as error:
        raise SystemExit(f"reconciliation failed: {error}") from error
