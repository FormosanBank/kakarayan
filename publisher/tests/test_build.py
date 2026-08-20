from __future__ import annotations

import hashlib
import io
import json
import sqlite3
import subprocess
import zipfile
from contextlib import closing
from pathlib import Path

import pyarrow as pa
import pyarrow.parquet as pq
import pytest
from jsonschema import ValidationError
from openpyxl import load_workbook

from publisher import PUBLIC_DOWNLOAD_PATHS
from publisher.build import BuildError, Source, _release_id, build_release
from publisher.cldf_export import write_cldf_package
from publisher.format_aligned import write_aligned_package
from publisher.verify_release import verify_release


def _tree_checksums(root: Path) -> dict[str, str]:
    return {
        path.relative_to(root).as_posix(): hashlib.sha256(path.read_bytes()).hexdigest()
        for path in root.rglob("*")
        if path.is_file()
    }


def test_release_identity_includes_source_and_application_revisions() -> None:
    source = Source(
        repository="FormosanBank/FormosanBank",
        commit="1" * 40,
        committed_at="2024-01-02T03:04:05Z",
    )

    assert _release_id(source, "2" * 40) == "fb-20240102-111111222222"
    assert _release_id(source, "3" * 40) == "fb-20240102-111111333333"


def test_fixture_release_is_valid_and_deterministic(public_repo: Path, tmp_path: Path) -> None:
    first = build_release(public_repo, tmp_path / "one")
    second = build_release(public_repo, tmp_path / "two")

    assert first.release_id.startswith("fb-20240102-")
    manifest = json.loads((first.output / "release-manifest.json").read_text(encoding="utf-8"))
    assert first.release_id.endswith(
        f"{first.source.commit[:6]}{manifest['kakarayan']['commit'][:6]}"
    )
    assert first.counts["texts"] == 1
    assert first.counts["tokens"] == 4
    assert first.warnings == ()
    assert _tree_checksums(first.output) == _tree_checksums(second.output)

    catalog = json.loads((first.output / "catalog.json").read_text(encoding="utf-8"))
    assert catalog["source"]["commit"] == first.source.commit
    assert catalog["corpora"][0]["name"] == "TestCorpus"
    assert catalog["corpora"][0]["rights_id"] == "rights_testcorpus"
    assert catalog["corpora"][0]["citation_count"] == 1
    amis = next(row for row in catalog["languages"] if row["name"] == "Amis")
    assert amis["counts"]["sentences"] == 2
    assert amis["dialects"] == ["Xiuguluan"]
    assert "audio" in amis["capabilities"]
    orthography = json.loads(
        (first.output / "api" / "v1" / "orthography.json").read_text(encoding="utf-8")
    )["data"]
    assert orthography["tables"][0]["language"] == "Amis"
    assert orthography["tables"][0]["rules"][1]["outputs"]["Xiuguluan"] == "o"
    content = json.loads(
        (first.output / "api" / "v1" / "content.json").read_text(encoding="utf-8")
    )["data"]
    assert content == {"schema_version": "1.0.0", "entries": []}
    downloads = json.loads(
        (first.output / "api" / "v1" / "downloads.json").read_text(encoding="utf-8")
    )["data"]
    assert downloads["release_id"] == first.release_id
    assert {item["path"] for item in downloads["artifacts"]} == (
        set(PUBLIC_DOWNLOAD_PATHS) - {"formosanbank.sqlite.gz"}
    )
    assert all(item["publishable"] for item in downloads["artifacts"])
    assert all(not item["blocked_reasons"] for item in downloads["artifacts"])
    assert all(item["format"] for item in downloads["artifacts"])
    assert all(item["language_ids"] == ["lang_amis"] for item in downloads["artifacts"])
    assert all(item["corpus_ids"] == ["corpus_testcorpus"] for item in downloads["artifacts"])
    assert all(item["tiers"] for item in downloads["artifacts"])
    assert all("/canonical/" not in item["path"] for item in downloads["artifacts"])
    assert all("/jsonl/" not in item["path"] for item in downloads["artifacts"])

    with zipfile.ZipFile(first.output / "prepared" / "parquet-tables.zip") as archive:
        assert pq.read_table(pa.BufferReader(archive.read("tokens.parquet"))).num_rows == 4
    workbook = load_workbook(
        first.output / "prepared" / "formosanbank.xlsx",
        read_only=True,
    )
    assert "README" in workbook.sheetnames
    assert "RIGHTS" in workbook.sheetnames
    workbook.close()
    canonical = next((first.output / "prepared" / "canonical").glob("*.zip"))
    with zipfile.ZipFile(canonical) as archive:
        source_path = "Corpora/TestCorpus/XML/fixture.xml"
        assert archive.read(source_path) == (public_repo / source_path).read_bytes()
        assert "rights.json" in archive.namelist()
    with zipfile.ZipFile(first.output / "prepared" / "time-aligned.zip") as archive:
        alignment = json.loads(archive.read("alignments.jsonl"))
        assert alignment["cues"][0]["sentence_id"].startswith("sentence_")
        assert alignment["cues"][0]["start_ms"] == 1250
        manifest = json.loads(archive.read("manifest.json"))
        assert manifest["media_groups"] == 1
        assert manifest["single_cue_groups"] == 1
        assert manifest["interchange_groups"] == 0
        assert "README.txt" in archive.namelist()
        assert "rights.json" in archive.namelist()
    with zipfile.ZipFile(first.output / "prepared" / "formosanbank-cldf.zip") as archive:
        assert "Generic-metadata.json" in archive.namelist()
        assert "README.txt" in archive.namelist()
        assert "rights.json" in archive.namelist()
    with zipfile.ZipFile(first.output / "prepared" / "hierarchical-jsonl.zip") as outer:
        [partition] = [name for name in outer.namelist() if name.endswith(".zip")]
        with zipfile.ZipFile(io.BytesIO(outer.read(partition))) as archive:
            assert archive.namelist() == [
                "README.txt",
                "data-dictionary.json",
                "part-0000.jsonl",
                "rights.json",
            ]

    with closing(sqlite3.connect(first.output / "formosanbank.sqlite")) as database:
        assert database.execute("PRAGMA integrity_check").fetchone() == ("ok",)
        assert database.execute("SELECT COUNT(*) FROM translations").fetchone() == (3,)
        embedded_meta = json.loads(
            database.execute(
                "SELECT value_json FROM publication_metadata WHERE key = 'meta'"
            ).fetchone()[0]
        )
        assert embedded_meta["release_id"] == first.release_id
        assert embedded_meta["kakarayan"]["commit"]


def test_output_directory_must_be_empty(public_repo: Path, tmp_path: Path) -> None:
    output = tmp_path / "output"
    output.mkdir()
    (output / "keep.txt").write_text("do not overwrite", encoding="utf-8")
    with pytest.raises(BuildError, match="absent or empty"):
        build_release(public_repo, output)


def test_source_commit_must_match(public_repo: Path, tmp_path: Path) -> None:
    with pytest.raises(BuildError, match="expected"):
        build_release(public_repo, tmp_path / "output", expected_commit="0" * 40)


def test_application_commit_must_match(public_repo: Path, tmp_path: Path) -> None:
    with pytest.raises(BuildError, match="Kakarayan HEAD"):
        build_release(public_repo, tmp_path / "output", application_commit="0" * 40)


def test_model_catalog_is_validated_before_corpus_projection(
    public_repo: Path,
    tmp_path: Path,
) -> None:
    output = tmp_path / "output"
    incomplete_catalog: dict[str, object] = {
        "schema_version": "1.0.0",
        "generated_at": "2024-01-02T03:04:05Z",
        "provider": "Hugging Face",
        "models": [
            {
                "id": "model_formosanbank_example",
                "repository": "FormosanBank/example",
                "task": "translation",
                "url": "https://huggingface.co/FormosanBank/example",
                "license": "unknown",
                "languages": ["ami"],
                "limitations": "Machine output.",
            }
        ],
        "services": [],
    }

    with pytest.raises(ValidationError, match="framework"):
        build_release(public_repo, output, model_catalog=incomplete_catalog)

    assert not (output / "tables").exists()


def test_database_can_be_packaged_for_github_releases(
    public_repo: Path,
    tmp_path: Path,
) -> None:
    release = build_release(
        public_repo,
        tmp_path / "release",
        compress_database=True,
        release_only=True,
    )
    assert not (release.output / "formosanbank.sqlite").exists()
    compressed = release.output / "formosanbank.sqlite.gz"
    assert compressed.read_bytes()[:2] == b"\x1f\x8b"
    manifest = verify_release(release.output, max_artifact_bytes=2 * 1024**3)
    artifact = next(
        item for item in manifest["artifacts"] if item["path"] == "formosanbank.sqlite.gz"
    )
    assert artifact["compression"] == "gzip"
    assert artifact["content_media_type"] == "application/vnd.sqlite3"
    assert artifact["asset_name"] == "formosanbank.sqlite.gz"
    assert artifact["download_url"].endswith(f"/data-{release.release_id}/formosanbank.sqlite.gz")
    assert not (release.output / "api").exists()
    assert (release.output / "site-metadata.zip").is_file()
    with zipfile.ZipFile(release.output / "site-metadata.zip") as archive:
        assert "v1/meta.json" in archive.namelist()
        assert not any(name.startswith("v1/search/") for name in archive.namelist())
    assert not (release.output / "tables").exists()
    asset_names = [item["asset_name"] for item in manifest["artifacts"]]
    assert len(asset_names) == len(set(asset_names))
    assert len(asset_names) == 12
    assert not (release.output / "prepared" / "canonical").exists()
    assert not (release.output / "prepared" / "jsonl").exists()
    with zipfile.ZipFile(release.output / "prepared" / "hierarchical-jsonl.zip") as archive:
        assert any(name.endswith(".zip") for name in archive.namelist())
    assert not any(
        (release.output / name).exists()
        for name in ("catalog.json", "models.json", "orthography.json", "rights.json")
    )


def test_multi_cue_media_gets_interchange_files(public_repo: Path, tmp_path: Path) -> None:
    release = build_release(public_repo, tmp_path / "release", include_prepared=False)
    database = release.output / "formosanbank.sqlite"
    with closing(sqlite3.connect(database)) as connection:
        sentence_ids = [
            row[0] for row in connection.execute("SELECT id FROM sentences ORDER BY id")
        ]
        connection.execute(
            """
            INSERT INTO audio (
              id, owner_type, owner_id, position, file, url, start, end,
              start_raw, end_raw, source, duration, availability_status, attributes_json
            ) VALUES (?, 'sentence', ?, 0, 'sentence.wav', NULL, 4.0, 5.0,
                      '4.0', '5.0', NULL, 1.0, 'unknown', '{}')
            """,
            ("audio_second_cue", sentence_ids[1]),
        )
        connection.commit()

    package = tmp_path / "time-aligned.zip"
    counts = write_aligned_package(database, package, release.release_id, {"entries": []})

    assert counts == {
        "media_groups": 1,
        "single_cue_groups": 0,
        "interchange_groups": 1,
        "textgrids": 1,
        "files": 10,
    }
    with zipfile.ZipFile(package) as archive:
        names = archive.namelist()
        assert any(name.endswith(".eaf") for name in names)
        assert any(name.endswith(".TextGrid") for name in names)
        assert any(name.endswith(".source.vtt") for name in names)
        assert any(name.endswith(".translation.srt") for name in names)


def test_cldf_streams_and_validates_large_source_fields(
    public_repo: Path,
    tmp_path: Path,
) -> None:
    release = build_release(
        public_repo,
        tmp_path / "release",
        include_prepared=False,
    )
    database = release.output / "formosanbank.sqlite"
    large_translation = "large field " * 15_000
    with closing(sqlite3.connect(database)) as connection:
        connection.execute(
            """
            UPDATE translations
            SET text = ?
            WHERE rowid = (SELECT MIN(rowid) FROM translations)
            """,
            (large_translation,),
        )
        connection.commit()
    rights = json.loads((release.output / "rights.json").read_text(encoding="utf-8"))
    counts = write_cldf_package(
        database,
        tmp_path / "large-cldf.zip",
        release_id=release.release_id,
        source_commit=release.source.commit,
        rights=rights,
    )
    assert counts["examples"] == 2
    with zipfile.ZipFile(tmp_path / "large-cldf.zip") as archive:
        assert large_translation.encode() in archive.read("examples.csv")


def test_cldf_includes_babuza_favorlang_language_reference(
    public_repo: Path,
    tmp_path: Path,
) -> None:
    [source] = public_repo.glob("Corpora/*/XML/*.xml")
    xml = source.read_text(encoding="utf-8")
    source.write_text(
        xml.replace('xml:lang="ami"', 'xml:lang="bzg"').replace(
            'dialect="Xiuguluan"', 'dialect="Favorlang"'
        ),
        encoding="utf-8",
    )
    subprocess.run(
        ["git", "-C", str(public_repo), "commit", "-am", "Use Babuza-Favorlang fixture"],
        check=True,
        capture_output=True,
    )
    release = build_release(
        public_repo,
        tmp_path / "release",
        include_prepared=False,
    )
    rights = json.loads((release.output / "rights.json").read_text(encoding="utf-8"))
    output = tmp_path / "babuza-favorlang-cldf.zip"
    counts = write_cldf_package(
        release.output / "formosanbank.sqlite",
        output,
        release_id=release.release_id,
        source_commit=release.source.commit,
        rights=rights,
    )

    assert counts["languages"] == 1
    with zipfile.ZipFile(output) as archive:
        languages = archive.read("languages.csv").decode("utf-8")
        examples = archive.read("examples.csv").decode("utf-8")
    assert "lang_babuza_favorlang,Babuza-Favorlang,,,,,bzg" in languages
    assert "lang_babuza_favorlang" in examples
